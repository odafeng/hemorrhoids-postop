#!/usr/bin/env python3
"""把研究人員 Dashboard 的完整備份填進 CRF 工作簿的可推導欄位。

資料庫是 source of truth；CRF 只承載資料庫生不出來的東西——簽名、納排條件、
補助費簽收、臨床判讀。某一欄是不是自動欄，看它有沒有出現在 SHEETS[...]['auto'] 裡；
沒出現的欄位這支腳本一個字都不會寫。

決策背景見 docs/adr/0002-crf-derived-from-db-export.md，
欄位分類見 docs/superpowers/specs/2026-08-29-crf-export-design.md。

用法：
    /usr/local/bin/python3 crf_fill.py [full_backup_YYYY-MM-DD.json]

不給路徑時取 ~/Downloads 裡最新的一份。這支腳本不連資料庫，也不需要金鑰。
"""
import json
import shutil
import sys
from copy import copy
from datetime import date, datetime
from pathlib import Path

import openpyxl

CRF_PATH = Path(__file__).resolve().parents[2] / '收案文件' / '個案報告表_CRF紀錄.xlsx'
TEST_PREFIX = 'TEST-'


class CrfError(Exception):
    """中止用。訊息會直接印給操作者看，所以寫人話。"""


def norm_date(value):
    """ISO 字串、datetime、date 一律正規化成 YYYY-MM-DD。

    對位鍵的兩端格式不同：JSON 給字串，Excel 給 datetime。不先攤平的話
    每次執行都會判定成新列，一路往下疊重複資料。
    """
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)[:10]


def load_backup(path=None, search_dir=None):
    if path:
        return json.loads(Path(path).read_text(encoding='utf-8'))
    search_dir = search_dir or Path.home() / 'Downloads'
    candidates = sorted(search_dir.glob('full_backup_*.json'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise CrfError(
            f'在 {search_dir} 找不到 full_backup_*.json。\n'
            '請先到研究人員 Dashboard 按「完整備份」下載一份。'
        )
    return json.loads(candidates[0].read_text(encoding='utf-8'))


def check_coverage(backup):
    """回傳缺手術記錄的 study_id。

    surgical_records 的 RLS 是 researcher_read_own_surgeon：非 PI 帳號只讀得到
    自己主刀的列，其餘不會報錯、只是不在備份裡。那些列填進 CRF 就是一片空白，
    看不出來是漏抓還是本來就沒有，所以寧可整支中止。
    """
    subjects = {p['study_id'] for p in backup.get('patients', [])
                if not p['study_id'].startswith(TEST_PREFIX)}
    have = {r['study_id'] for r in backup.get('surgical_records', [])}
    return sorted(subjects - have)


def header_map(ws, header_row):
    return {cell.value: cell.column
            for cell in ws[header_row] if cell.value is not None}


def backup_workbook(path):
    dest = path.with_name(f"{path.stem}.bak-{datetime.now():%Y%m%d}{path.suffix}")
    shutil.copy2(path, dest)
    return dest


# surgical_records 存的是代碼（closed、ligasure、LMGA），CRF「選項清單」分頁宣告的
# 允許值是人看的標籤。症狀類欄位不需要這層：DB 存的就是「少量」「未排」這種短式，
# 與選項清單的首詞和既有手填列都一致，原值寫入即可。
#
# 代碼集合由 test_crf_fill 對 src/pages/SurgicalRecord.jsx 比對釘住：App 加了新選項
# 而這裡沒跟上時測試會紅，而不是安靜地寫入空白。
PROCEDURE_LABEL = {
    'hemorrhoidectomy': 'Hemorrhoidectomy',
    'laser_hemorrhoidoplasty': 'Laser hemorrhoidoplasty',
}
SUBTYPE_LABEL = {
    'open': 'Open（傷口不縫合）',
    'closed': 'Closed（Ferguson，傷口完全縫合）',
    'semi_open': 'Semi-open（部分縫合，中央開放）',
    'semi_closed': 'Semi-closed（大部分縫合，末端開放）',
}
ANESTHESIA_LABEL = {
    'IVGA': 'IVGA（靜脈全身麻醉）',
    'LMGA': 'LMGA（喉罩全身麻醉）',
    'SA': 'SA（脊椎麻醉）',
    'LA': 'LA（局部麻醉）',
}
ENERGY_LABEL = {
    'ligasure': 'LigaSure',
    'powerseal': 'Powerseal',
    'harmonic': 'Harmonic',
}
STATUS_LABEL = {
    'active': '追蹤中',
    'completed': '已完成',
    'withdrawn': '已退出',
}
# 與 src/utils/hooks.js 的 ALERT_DISPLAY 標題一致，也與 CRF 既有手填列一致。
ALERT_TYPE_LABEL = {
    'high_pain': '持續性高度疼痛',
    'ascending_pain': '疼痛逐日上升',
    'persistent_bleeding': '持續性出血',
    'blood_clot': '出血伴隨血塊',
    'no_bowel': '超過3天未排便',
    'fever': '發燒',
    'urinary_retention': '完全尿不出來',
    'urinary_difficulty': '排尿困難',
    'incontinence': '肛門失禁',
    'soiling': '持續滲便',
}
ALERT_LEVEL_LABEL = {
    'danger': '危險（danger）',
    'warning': '警告（warning）',
    'info': '提示（info）',
}


def _label(mapping, code):
    """查不到就原值送出，讓沒對到的代碼在 CRF 上看得見，而不是變成空白。"""
    if code is None or code == '':
        return None
    return mapping.get(code, code)


def _energy(devices):
    if not devices:
        return '無'
    return '、'.join(_label(ENERGY_LABEL, d) for d in devices)


def _is_date_column(name):
    return '日期' in name or '時間' in name


def _row_key(ws, row, idx, key_cols):
    return tuple(
        norm_date(ws.cell(row=row, column=idx[name]).value) if _is_date_column(name)
        else ws.cell(row=row, column=idx[name]).value
        for name in key_cols
    )


def _first_blank_row(ws, header_row, idx, key_cols):
    col = idx[key_cols[0]]
    row = header_row + 1
    while ws.cell(row=row, column=col).value not in (None, ''):
        row += 1
    return row


def upsert_sheet(ws, header_row, key_cols, records, auto_map):
    """把 records 寫進 ws：auto_map 裡的欄位覆寫，其餘一律不動。

    對位靠 key_cols 組成的鍵，不靠列號。列序會因為補進較早日期的資料而改變，
    照列號寫會把手填的備註接到別人的資料上——而那種錯不會有任何一端報錯。
    """
    idx = header_map(ws, header_row)
    missing = [c for c in list(key_cols) + list(auto_map) if c not in idx]
    if missing:
        raise CrfError(f'「{ws.title}」找不到欄位：{"、".join(missing)}')

    existing = {}
    for row in range(header_row + 1, ws.max_row + 1):
        key = _row_key(ws, row, idx, key_cols)
        if key[0] not in (None, ''):
            existing[key] = row

    written = 0
    for rec in records:
        key = tuple(
            norm_date(auto_map[name](rec)) if _is_date_column(name) else auto_map[name](rec)
            for name in key_cols
        )
        row = existing.get(key)
        if row is None:
            row = _first_blank_row(ws, header_row, idx, key_cols)
            if row > header_row + 1:
                # 整列複製樣式。只設 font 會漏掉 alignment / border / fill，
                # 新列就停在空白列的「待填」外觀（研究日誌/2026-08-13.yaml）。
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col)._style = copy(
                        ws.cell(row=row - 1, column=col)._style)
            existing[key] = row
        for name, fn in auto_map.items():
            ws.cell(row=row, column=idx[name]).value = fn(rec)
        written += 1
    return written


def build_context(backup):
    """把備份攤平成每個分頁要的記錄。跨表資料先併好，欄位 lambda 裡不再查表。"""
    by_id = {p['study_id']: p for p in backup.get('patients', [])
             if not p['study_id'].startswith(TEST_PREFIX)}
    adh = {a['study_id']: a for a in backup.get('adherence_summary', [])}
    surg = {s['study_id']: s for s in backup.get('surgical_records', [])}
    surveys = {s['study_id']: s for s in backup.get('usability_surveys', [])}

    reports = [r for r in backup.get('symptom_reports', []) if r['study_id'] in by_id]
    alerts = [a for a in backup.get('alerts', []) if a['study_id'] in by_id]
    hcu = [h for h in backup.get('healthcare_utilization', []) if h['study_id'] in by_id]
    chats = [c for c in backup.get('ai_chat_logs', []) if c.get('study_id') in by_id]

    def count(rows, sid):
        return sum(1 for r in rows if r.get('study_id') == sid)

    def pod_of(sid, when):
        surgery = norm_date(by_id[sid].get('surgery_date'))
        when = norm_date(when)
        if not surgery or not when:
            return None
        return (date.fromisoformat(when) - date.fromisoformat(surgery)).days

    overview = []
    for seq, (sid, p) in enumerate(
            sorted(by_id.items(), key=lambda kv: norm_date(kv[1].get('created_at')) or ''),
            start=1):
        overview.append({
            'seq': seq, 'study_id': sid, 'patient': p,
            'adherence': adh.get(sid, {}), 'surgical': surg.get(sid, {}),
            'survey': surveys.get(sid),
            'n_reports': count(reports, sid),
            'n_alerts': count(alerts, sid),
            'n_unacked': sum(1 for a in alerts
                             if a['study_id'] == sid and not a.get('acknowledged')),
            'n_chats': count(chats, sid),
        })

    return {
        'overview': overview,
        'reports': reports,
        'alerts': [dict(a, _pod=pod_of(a['study_id'], a.get('triggered_at'))) for a in alerts],
        'hcu': hcu,
        'closed': [o for o in overview
                   if o['patient'].get('study_status') in ('completed', 'withdrawn')],
    }


# 一個欄位是不是自動欄，就看它在不在這裡。手填欄（簽名、納排條件、補助費、
# 警示處理、資料澄清）與公式欄（POD（今日）、依從率）刻意不列，腳本因此碰不到。
SHEETS = {
    '個案總覽': {
        'header_row': 4, 'key': ('Study ID',), 'source': 'overview',
        'auto': {
            '序號': lambda r: r['seq'],
            'Study ID': lambda r: r['study_id'],
            '主刀醫師': lambda r: r['patient'].get('surgeon_id'),
            '手術日期': lambda r: norm_date(r['patient'].get('surgery_date')),
            '收案日期': lambda r: norm_date(r['patient'].get('created_at')),
            '實際回報數': lambda r: r['n_reports'],
            '應回報數': lambda r: r['adherence'].get('expected_reports'),
            'AI 衛教使用次數': lambda r: r['n_chats'],
            '警示總數': lambda r: r['n_alerts'],
            '未確認警示': lambda r: r['n_unacked'],
            '可用性問卷': lambda r: '已完成' if r['survey'] else '未完成',
            '收案狀態': lambda r: _label(STATUS_LABEL, r['patient'].get('study_status')),
        },
    },
    '表單一_收案登記': {
        'header_row': 4, 'key': ('Study ID',), 'source': 'overview',
        'auto': {
            'Study ID': lambda r: r['study_id'],
            '收案日期': lambda r: norm_date(r['patient'].get('created_at')),
            '痔瘡分級': lambda r: r['surgical'].get('hemorrhoid_grade'),
            '手術日期': lambda r: norm_date(r['patient'].get('surgery_date')),
            '術式': lambda r: _label(PROCEDURE_LABEL, r['surgical'].get('procedure_type')),
            '縫合方式': lambda r: _label(SUBTYPE_LABEL, r['surgical'].get('hemorrhoidectomy_subtype')),
            '能量器械': lambda r: _energy(r['surgical'].get('energy_device')),
            '麻醉方式': lambda r: _label(ANESTHESIA_LABEL, r['surgical'].get('anesthesia_type')),
            '主刀醫師': lambda r: r['patient'].get('surgeon_id'),
            '同意書簽署日': lambda r: norm_date(r['patient'].get('consent_date')),
            # patients 有這一列就代表註冊完成。不讀 app_activated：那個欄位有 DEFAULT
            # 沒有寫入端，全部受試者都是 false，拿它判定會整欄填成「否」。
            'App 註冊完成': lambda r: '是',
        },
    },
    '表單二_每日症狀回報': {
        'header_row': 5, 'key': ('Study ID', '回報日期'), 'source': 'reports',
        'auto': {
            'Study ID': lambda r: r['study_id'],
            '回報日期': lambda r: norm_date(r.get('report_date')),
            'POD': lambda r: r.get('pod'),
            '疼痛 NRS': lambda r: r.get('pain_nrs'),
            '出血程度': lambda r: r.get('bleeding'),
            '排便狀況': lambda r: r.get('bowel'),
            '肛門控制': lambda r: r.get('continence'),
            '發燒': lambda r: '是（體溫 ≥ 38°C）' if r.get('fever') else '否',
            '排尿狀況': lambda r: r.get('urinary'),
            # wound 是逗號分隔字串，不是陣列（schemaContract.formatWound 靠 split(',')）。
            # 既有手填列也是「異物感,腫脹」這個格式，原值寫入。
            '傷口狀況（可複選）': lambda r: r.get('wound'),
            '資料來源': lambda r: r.get('report_source'),
        },
    },
    '表單三_警示處理紀錄': {
        'header_row': 4, 'key': ('Study ID', '警示日期'), 'source': 'alerts',
        'auto': {
            'Study ID': lambda r: r['study_id'],
            '警示日期': lambda r: norm_date(r.get('triggered_at')),
            'POD': lambda r: r.get('_pod'),
            '警示類型': lambda r: _label(ALERT_TYPE_LABEL, r.get('alert_type')),
            '警示等級': lambda r: _label(ALERT_LEVEL_LABEL, r.get('alert_level')),
        },
    },
    '表單四_醫療利用紀錄': {
        'header_row': 4, 'key': ('Study ID', '就醫日期'), 'source': 'hcu',
        'auto': {
            'Study ID': lambda r: r['study_id'],
            '就醫日期': lambda r: norm_date(r.get('event_date')),
            'POD': lambda r: r.get('pod_at_event'),
            '就醫類型': lambda r: r.get('event_type'),
            '就醫原因': lambda r: r.get('reason'),
        },
    },
    '表單六_結案退出': {
        'header_row': 4, 'key': ('Study ID',), 'source': 'closed',
        'auto': {
            'Study ID': lambda r: r['study_id'],
            '結案狀態': lambda r: _label(STATUS_LABEL, r['patient'].get('study_status')),
            '結案/退出日期': lambda r: norm_date(r['patient'].get('completed_at')),
            'POD': lambda r: r['adherence'].get('max_pod'),
            '總回報次數': lambda r: r['n_reports'],
            '預期次數': lambda r: r['adherence'].get('expected_reports'),
            'AI 衛教使用次數': lambda r: r['n_chats'],
            '可用性問卷': lambda r: '已完成' if r['survey'] else '未完成',
        },
    },
}


def run(backup, crf_path=CRF_PATH):
    """填入自動欄。中止時完全沒有副作用，包含不留快照。"""
    missing = check_coverage(backup)
    if missing:
        raise CrfError(
            '這幾例沒有手術記錄，已中止，未寫入任何內容：\n'
            f'  {"、".join(missing)}\n'
            'surgical_records 的 RLS 只讓研究人員看到自己主刀的列。若備份不是用 PI 帳號\n'
            '下載的，請改用 PI 帳號重新下載；若確實是主刀醫師還沒登錄，請先補登手術記錄。'
        )

    snapshot = backup_workbook(crf_path)
    ctx = build_context(backup)
    wb = openpyxl.load_workbook(crf_path)
    summary = []
    for name, spec in SHEETS.items():
        if name not in wb.sheetnames:
            raise CrfError(f'工作簿裡沒有分頁「{name}」')
        n = upsert_sheet(wb[name], spec['header_row'], spec['key'],
                         ctx[spec['source']], spec['auto'])
        summary.append(f'  {name}：{n} 列')
    wb.save(crf_path)
    return snapshot, summary


def main(argv):
    try:
        backup = load_backup(argv[1] if len(argv) > 1 else None)
        snapshot, summary = run(backup)
    except CrfError as err:
        print(f'中止：{err}', file=sys.stderr)
        return 1
    print(f'快照：{snapshot.name}')
    print('\n'.join(summary))
    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv))
