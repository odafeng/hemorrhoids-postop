"""crf_fill 的單元測試。

不在 CI 裡（ci.yml 不裝 openpyxl），手動執行：
    cd prototype/scripts && /usr/local/bin/python3 -m unittest test_crf_fill -v

系統的 /usr/bin/python3 沒有 openpyxl，要用 /usr/local/bin/python3。
"""
import json
import os
import pathlib
import re
import tempfile
import unittest
from datetime import date, datetime

import openpyxl

import crf_fill


class TestNormDate(unittest.TestCase):
    """對位鍵一邊來自 JSON（字串）、一邊來自 Excel（datetime）。
    不正規化就永遠對不上，每次執行都會新增一份重複列。"""

    def test_accepts_iso_string_with_time(self):
        self.assertEqual(crf_fill.norm_date('2026-08-11T09:30:00+00:00'), '2026-08-11')

    def test_accepts_plain_date_string(self):
        self.assertEqual(crf_fill.norm_date('2026-08-11'), '2026-08-11')

    def test_accepts_excel_datetime(self):
        self.assertEqual(crf_fill.norm_date(datetime(2026, 8, 11, 0, 0)), '2026-08-11')

    def test_accepts_date(self):
        self.assertEqual(crf_fill.norm_date(date(2026, 8, 11)), '2026-08-11')

    def test_blank_is_none(self):
        self.assertIsNone(crf_fill.norm_date(None))
        self.assertIsNone(crf_fill.norm_date(''))


class TestCoverage(unittest.TestCase):
    def _backup(self, patients, surgical):
        return {'patients': patients, 'surgical_records': surgical}

    def test_all_covered_returns_empty(self):
        b = self._backup(
            [{'study_id': 'AAA-001'}, {'study_id': 'AAA-002'}],
            [{'study_id': 'AAA-001'}, {'study_id': 'AAA-002'}],
        )
        self.assertEqual(crf_fill.check_coverage(b), [])

    def test_missing_surgical_record_is_reported(self):
        b = self._backup(
            [{'study_id': 'AAA-001'}, {'study_id': 'AAA-002'}],
            [{'study_id': 'AAA-001'}],
        )
        self.assertEqual(crf_fill.check_coverage(b), ['AAA-002'])

    def test_test_account_is_exempt(self):
        b = self._backup(
            [{'study_id': 'TEST-001'}, {'study_id': 'AAA-001'}],
            [{'study_id': 'AAA-001'}],
        )
        self.assertEqual(crf_fill.check_coverage(b), [])


class TestHeaderMap(unittest.TestCase):
    def test_maps_names_to_one_based_columns(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A4'], ws['B4'], ws['C4'] = 'Study ID', '收案日期', '年齡'
        self.assertEqual(crf_fill.header_map(ws, 4),
                         {'Study ID': 1, '收案日期': 2, '年齡': 3})


class TestLoadBackup(unittest.TestCase):
    def test_picks_newest_when_no_path_given(self):
        with tempfile.TemporaryDirectory() as d:
            old = pathlib.Path(d) / 'full_backup_2026-08-01.json'
            new = pathlib.Path(d) / 'full_backup_2026-08-29.json'
            old.write_text(json.dumps({'patients': [{'study_id': 'OLD'}]}))
            new.write_text(json.dumps({'patients': [{'study_id': 'NEW'}]}))
            os.utime(old, (1, 1))
            got = crf_fill.load_backup(search_dir=pathlib.Path(d))
            self.assertEqual(got['patients'][0]['study_id'], 'NEW')

    def test_no_backup_found_raises(self):
        with tempfile.TemporaryDirectory() as d:
            with self.assertRaises(crf_fill.CrfError):
                crf_fill.load_backup(search_dir=pathlib.Path(d))


class TestBackupWorkbook(unittest.TestCase):
    def test_snapshot_keeps_the_gitignored_prefix(self):
        with tempfile.TemporaryDirectory() as d:
            src = pathlib.Path(d) / '個案報告表_CRF紀錄.xlsx'
            openpyxl.Workbook().save(src)
            dest = crf_fill.backup_workbook(src)
            self.assertTrue(dest.exists())
            # .gitignore:40 的 glob 是 收案文件/個案報告表*.xlsx
            self.assertTrue(dest.name.startswith('個案報告表'))
            self.assertTrue(dest.name.endswith('.xlsx'))


class TestUpsert(unittest.TestCase):
    AUTO = {
        'Study ID': lambda r: r['study_id'],
        '回報日期': lambda r: crf_fill.norm_date(r['report_date']),
        '疼痛 NRS': lambda r: r['pain_nrs'],
    }
    KEY = ('Study ID', '回報日期')

    def _sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        for col, name in enumerate(['Study ID', '回報日期', '疼痛 NRS', '備註'], start=1):
            ws.cell(row=5, column=col).value = name
        return ws

    def test_inserts_new_rows(self):
        ws = self._sheet()
        crf_fill.upsert_sheet(ws, 5, self.KEY, [
            {'study_id': 'AAA-001', 'report_date': '2026-08-11', 'pain_nrs': 3},
        ], self.AUTO)
        self.assertEqual(ws.cell(row=6, column=1).value, 'AAA-001')
        self.assertEqual(ws.cell(row=6, column=3).value, 3)

    def test_manual_column_survives_rerun(self):
        ws = self._sheet()
        rec = {'study_id': 'AAA-001', 'report_date': '2026-08-11', 'pain_nrs': 3}
        crf_fill.upsert_sheet(ws, 5, self.KEY, [rec], self.AUTO)
        ws.cell(row=6, column=4).value = '主持人判讀為傷口分泌物'
        rec['pain_nrs'] = 5
        crf_fill.upsert_sheet(ws, 5, self.KEY, [rec], self.AUTO)
        self.assertEqual(ws.cell(row=6, column=3).value, 5)
        self.assertEqual(ws.cell(row=6, column=4).value, '主持人判讀為傷口分泌物')

    def test_manual_column_follows_its_own_row_when_order_changes(self):
        """靠列號對位的實作在前兩個測試也會過，只有這裡會現形：
        插入一筆較早的日期讓列序改變，手填備註不能接到別人身上。"""
        ws = self._sheet()
        later = {'study_id': 'AAA-001', 'report_date': '2026-08-11', 'pain_nrs': 3}
        crf_fill.upsert_sheet(ws, 5, self.KEY, [later], self.AUTO)
        ws.cell(row=6, column=4).value = '屬於 08-11 的備註'

        earlier = {'study_id': 'AAA-001', 'report_date': '2026-08-09', 'pain_nrs': 7}
        crf_fill.upsert_sheet(ws, 5, self.KEY, [earlier, later], self.AUTO)

        rows = {crf_fill.norm_date(ws.cell(row=r, column=2).value): r for r in (6, 7)}
        self.assertEqual(ws.cell(row=rows['2026-08-11'], column=4).value, '屬於 08-11 的備註')
        self.assertIsNone(ws.cell(row=rows['2026-08-09'], column=4).value)

    def test_excel_datetime_key_matches_json_string(self):
        ws = self._sheet()
        ws.cell(row=6, column=1).value = 'AAA-001'
        ws.cell(row=6, column=2).value = datetime(2026, 8, 11)
        ws.cell(row=6, column=4).value = '既有備註'
        crf_fill.upsert_sheet(ws, 5, self.KEY, [
            {'study_id': 'AAA-001', 'report_date': '2026-08-11T00:00:00+00:00', 'pain_nrs': 9},
        ], self.AUTO)
        self.assertIsNone(ws.cell(row=7, column=1).value)  # 沒有新增重複列
        self.assertEqual(ws.cell(row=6, column=3).value, 9)
        self.assertEqual(ws.cell(row=6, column=4).value, '既有備註')

    def test_new_row_copies_style_from_previous_row(self):
        ws = self._sheet()
        crf_fill.upsert_sheet(ws, 5, self.KEY, [
            {'study_id': 'AAA-001', 'report_date': '2026-08-09', 'pain_nrs': 1},
        ], self.AUTO)
        ws.cell(row=6, column=1).font = openpyxl.styles.Font(bold=True, size=13)
        crf_fill.upsert_sheet(ws, 5, self.KEY, [
            {'study_id': 'AAA-001', 'report_date': '2026-08-09', 'pain_nrs': 1},
            {'study_id': 'AAA-002', 'report_date': '2026-08-09', 'pain_nrs': 2},
        ], self.AUTO)
        self.assertTrue(ws.cell(row=7, column=1).font.bold)
        self.assertEqual(ws.cell(row=7, column=1).font.size, 13)

    def test_missing_column_aborts(self):
        ws = self._sheet()
        with self.assertRaises(crf_fill.CrfError):
            crf_fill.upsert_sheet(ws, 5, self.KEY, [
                {'study_id': 'AAA-001', 'report_date': '2026-08-09', 'pain_nrs': 1},
            ], {**self.AUTO, '不存在的欄位': lambda r: 1})


class TestSheetMapping(unittest.TestCase):
    """自動欄與手填欄的界線就是這支腳本的整個安全性論證，所以直接測它。"""

    FORMULA_COLUMNS = {
        '個案總覽': ['POD（今日）', '依從率'],
        '表單六_結案退出': ['依從率'],
    }
    MANUAL_COLUMNS = {
        '個案總覽': ['備註'],
        '表單一_收案登記': ['年齡', '性別', 'BMI', '納入條件全符合', '排除條件皆無',
                            '補助費 NT$300', '研究人員簽名', '備註'],
        '表單二_每日症狀回報': ['備註'],
        '表單三_警示處理紀錄': ['處理方式', '處理結果', '是否為假陽性', '處理人員', '處理日期'],
        '表單四_醫療利用紀錄': ['處置內容', '記錄人員', '記錄日期'],
        '表單六_結案退出': ['退出原因', '備註', '研究人員簽名'],
    }

    def test_no_manual_column_is_ever_automated(self):
        for sheet, manual in self.MANUAL_COLUMNS.items():
            auto = set(crf_fill.SHEETS[sheet]['auto'])
            self.assertEqual(auto & set(manual), set(), f'{sheet} 把手填欄列成自動欄')

    def test_formula_columns_are_never_written(self):
        for sheet, formulas in self.FORMULA_COLUMNS.items():
            auto = set(crf_fill.SHEETS[sheet]['auto'])
            self.assertEqual(auto & set(formulas), set(),
                             f'{sheet} 會把公式覆寫成靜態值')

    def test_資料澄清註記_is_not_touched_at_all(self):
        self.assertNotIn('資料澄清註記', crf_fill.SHEETS)

    def test_app_registration_does_not_read_app_activated(self):
        """app_activated 有 DEFAULT 但沒有寫入端，全體受試者都是 false。
        拿它判定「App 註冊完成」會整欄填成「否」而且沒有一端報錯。"""
        src = pathlib.Path(crf_fill.__file__).read_text(encoding='utf-8')
        self.assertNotIn("get('app_activated')", src)

    def test_every_source_key_exists_in_context(self):
        ctx = crf_fill.build_context(_MINIMAL_BACKUP)
        for name, spec in crf_fill.SHEETS.items():
            self.assertIn(spec['source'], ctx, f'{name} 的 source 不在 context 裡')

    def test_adherence_comes_from_the_view_not_a_recount(self):
        """應回報數的單一定義是 fn_report_days()，v_adherence_summary 已經在用它。
        腳本自己重算就會出現三方不一致。"""
        ctx = crf_fill.build_context(_MINIMAL_BACKUP)
        rec = ctx['overview'][0]
        self.assertEqual(crf_fill.SHEETS['個案總覽']['auto']['應回報數'](rec), 9)
        # 實際回報數是真的數 symptom_reports，兩者不該相等
        self.assertEqual(crf_fill.SHEETS['個案總覽']['auto']['實際回報數'](rec), 1)


class TestSurgicalCodeCoverage(unittest.TestCase):
    """代碼標籤是 SurgicalRecord.jsx 的第二份副本。App 加了新選項而這裡沒跟上時，
    _label 會原值送出、CRF 上出現一個 raw code；這組測試讓它在那之前就轉紅。"""

    JSX = (pathlib.Path(crf_fill.__file__).resolve().parents[1]
           / 'src' / 'pages' / 'SurgicalRecord.jsx')

    def _codes_in(self, block_name):
        src = self.JSX.read_text(encoding='utf-8')
        start = src.index(f'const {block_name} = [')
        block = src[start:src.index('];', start)]
        return set(re.findall(r"\{\s*v:\s*'([^']+)'", block))

    def test_subtype_labels_cover_every_app_option(self):
        self.assertEqual(self._codes_in('HEM_SUBTYPES'), set(crf_fill.SUBTYPE_LABEL))

    def test_anesthesia_labels_cover_every_app_option(self):
        self.assertEqual(self._codes_in('ANESTHESIA_TYPES'), set(crf_fill.ANESTHESIA_LABEL))

    def test_energy_labels_cover_every_app_option(self):
        self.assertEqual(self._codes_in('ENERGY_DEVICES'), set(crf_fill.ENERGY_LABEL))

    def test_procedure_types_match_the_app(self):
        src = self.JSX.read_text(encoding='utf-8')
        for code in crf_fill.PROCEDURE_LABEL:
            self.assertIn(f"setProcedureType('{code}')", src)

    def test_no_energy_device_reads_as_none_not_blank(self):
        self.assertEqual(crf_fill._energy([]), '無')
        self.assertEqual(crf_fill._energy(None), '無')
        self.assertEqual(crf_fill._energy(['ligasure']), 'LigaSure')

    def test_unmapped_code_survives_instead_of_vanishing(self):
        self.assertEqual(crf_fill._label(crf_fill.SUBTYPE_LABEL, 'brand_new'), 'brand_new')


_MINIMAL_BACKUP = {
    'patients': [{
        'study_id': 'AAA-001', 'surgery_date': '2026-08-01',
        'created_at': '2026-07-31T02:00:00+00:00',
        'consent_date': '2026-07-31T02:00:00+00:00',
        'surgeon_id': 'AAA', 'study_status': 'active',
    }],
    'surgical_records': [{
        'study_id': 'AAA-001', 'procedure_type': 'hemorrhoidectomy',
        'hemorrhoidectomy_subtype': 'closed', 'hemorrhoid_grade': 'III',
        'anesthesia_type': 'LMGA', 'energy_device': [],
    }],
    'symptom_reports': [{
        'study_id': 'AAA-001', 'report_date': '2026-08-01', 'pod': 0,
        'pain_nrs': 3, 'bleeding': '少量', 'bowel': '未排', 'fever': False,
        'urinary': '正常', 'continence': '正常', 'wound': '異物感,腫脹',
        'report_source': 'app',
    }],
    'alerts': [], 'ai_chat_logs': [], 'usability_surveys': [],
    'healthcare_utilization': [],
    'adherence_summary': [{'study_id': 'AAA-001', 'expected_reports': 9, 'max_pod': 0}],
}


if __name__ == '__main__':
    unittest.main()
