"""
Build the Study ID ↔ Patient Identity crosswalk Excel template.
PI-eyes-only document — store encrypted / password-protected.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "收案對照表"

# ── Palette ──
BLUE      = "0F4C75"
DARK_BLUE = "0A3550"
WHITE     = "FFFFFF"
LIGHT_BG  = "F2F7FB"
WARN_BG   = "FFF8E6"
LINE_CLR  = "C5D0DC"
ACCENT    = "3282B8"

header_font  = Font(name="Noto Sans TC", size=10, bold=True, color=WHITE)
header_fill  = PatternFill("solid", fgColor=BLUE)
sub_font     = Font(name="Noto Sans TC", size=9, color=DARK_BLUE)
mono_font    = Font(name="Menlo", size=10, bold=True, color=BLUE)
date_font    = Font(name="Menlo", size=9, color="444444")
note_font    = Font(name="Noto Sans TC", size=8, color="888888", italic=True)
warn_font    = Font(name="Noto Sans TC", size=9, bold=True, color="B8600A")
warn_fill    = PatternFill("solid", fgColor=WARN_BG)
stripe_fill  = PatternFill("solid", fgColor=LIGHT_BG)
thin_border  = Border(
    left=Side(style="thin", color=LINE_CLR),
    right=Side(style="thin", color=LINE_CLR),
    top=Side(style="thin", color=LINE_CLR),
    bottom=Side(style="thin", color=LINE_CLR),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Title rows ──
ws.merge_cells("A1:J1")
title_cell = ws["A1"]
title_cell.value = "研究編號 ↔ 病人身份對照表（PI 專用 · 機密）"
title_cell.font = Font(name="Noto Sans TC", size=14, bold=True, color=DARK_BLUE)
title_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:J2")
sub_cell = ws["A2"]
sub_cell.value = (
    "結合 AI 衛教之痔瘡術後數位症狀監測系統 · KSVGH CRS · "
    f"建立日期：{date.today().isoformat()}　"
    "⚠ 本檔案含個人可識別資訊，須加密保存，僅限 PI 存取"
)
sub_cell.font = Font(name="Noto Sans TC", size=8.5, color="777777")
sub_cell.alignment = left
ws.row_dimensions[2].height = 20

# Warning banner
ws.merge_cells("A3:J3")
warn_cell = ws["A3"]
warn_cell.value = "⚠ 本對照表為 IRB 規定之個資文件，請以加密方式保存（如 Excel 密碼保護 / 加密隨身碟），並僅限計畫主持人存取。"
warn_cell.font = warn_font
warn_cell.fill = warn_fill
warn_cell.alignment = left
warn_cell.border = thin_border
ws.row_dimensions[3].height = 24

# ── Column headers (row 5) ──
HEADERS = [
    ("A", "序號",         6),
    ("B", "Study ID",    14),
    ("C", "姓名",        12),
    ("D", "病歷號",      14),
    ("E", "手術日期",    13),
    ("F", "主刀醫師",    12),
    ("G", "聯絡電話",    15),
    ("H", "收案日期",    13),
    ("I", "收案狀態",    12),
    ("J", "備註",        20),
]

HEADER_ROW = 5
ws.row_dimensions[HEADER_ROW].height = 26

for col_letter, title, width in HEADERS:
    cell = ws[f"{col_letter}{HEADER_ROW}"]
    cell.value = title
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border
    ws.column_dimensions[col_letter].width = width

# ── Surgeon list for data validation ──
SURGEONS = {
    "HSF": "黃士峯", "HCW": "許詔文", "WJH": "王瑞和", "CPT": "朱炳騰",
    "WCC": "吳志謙", "LMH": "李明泓", "CYH": "陳禹勳", "FIH": "方翊軒",
}
surgeon_list = ",".join(f"{v}（{k}）" for k, v in SURGEONS.items())
surgeon_dv = DataValidation(type="list", formula1=f'"{surgeon_list}"', allow_blank=True)
surgeon_dv.error = "請從下拉選單選擇主刀醫師"
surgeon_dv.errorTitle = "無效的醫師"
ws.add_data_validation(surgeon_dv)

# Status validation
status_dv = DataValidation(
    type="list",
    formula1='"待收案,已註冊,已簽同意書,追蹤中,已完成,已退出"',
    allow_blank=True,
)
status_dv.error = "請從下拉選單選擇狀態"
ws.add_data_validation(status_dv)

# ── Pre-fill 50 rows (target enrollment) ──
DATA_START = 6
NUM_ROWS = 50

for i in range(NUM_ROWS):
    row = DATA_START + i
    ws.row_dimensions[row].height = 22
    is_stripe = i % 2 == 1

    for col_letter, _, _ in HEADERS:
        cell = ws[f"{col_letter}{row}"]
        cell.border = thin_border
        cell.alignment = center if col_letter in ("A", "B", "E", "F", "H", "I") else left
        if is_stripe:
            cell.fill = stripe_fill

    # 序號
    ws[f"A{row}"].value = i + 1
    ws[f"A{row}"].font = Font(name="Menlo", size=9, color="AAAAAA")

    # Study ID placeholder (mono font)
    ws[f"B{row}"].font = mono_font
    ws[f"B{row}"].number_format = "@"  # text

    # Date columns
    ws[f"E{row}"].font = date_font
    ws[f"E{row}"].number_format = "YYYY-MM-DD"
    ws[f"H{row}"].font = date_font
    ws[f"H{row}"].number_format = "YYYY-MM-DD"

    # Apply data validations
    surgeon_dv.add(f"F{row}")
    status_dv.add(f"I{row}")

    # Note column
    ws[f"J{row}"].font = note_font

# ── Example row (row 6) ──
ex = DATA_START
ws[f"B{ex}"].value = "HSF-001"
ws[f"C{ex}"].value = "（範例）王○明"
ws[f"C{ex}"].font = Font(name="Noto Sans TC", size=9, color="BBBBBB", italic=True)
ws[f"D{ex}"].value = "（範例）12345678"
ws[f"D{ex}"].font = Font(name="Menlo", size=9, color="BBBBBB", italic=True)
ws[f"E{ex}"].value = date(2026, 6, 4)
ws[f"F{ex}"].value = "黃士峯（HSF）"
ws[f"G{ex}"].value = "（範例）0912-345-678"
ws[f"G{ex}"].font = Font(name="Menlo", size=9, color="BBBBBB", italic=True)
ws[f"H{ex}"].value = date(2026, 6, 4)
ws[f"I{ex}"].value = "追蹤中"
ws[f"J{ex}"].value = "範例資料，請刪除後使用"
ws[f"J{ex}"].font = Font(name="Noto Sans TC", size=8, color="CCCCCC", italic=True)

# ── Footer notes ──
footer_row = DATA_START + NUM_ROWS + 1
ws.merge_cells(f"A{footer_row}:J{footer_row}")
fc = ws[f"A{footer_row}"]
fc.value = (
    "【保管規定】本對照表依 IRB 規定，僅限計畫主持人（PI）存取。"
    "須以加密方式保存（建議：Excel 檔案密碼保護 + 存放於加密磁碟）。"
    "研究結束後依院內規定保存後銷毀。去識別化資料保存 10 年。"
)
fc.font = Font(name="Noto Sans TC", size=8, color="999999")
fc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws.row_dimensions[footer_row].height = 30

# ── Freeze panes: header row stays visible ──
ws.freeze_panes = f"A{DATA_START}"

# ── Print settings ──
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"

# ── Save ──
out_path = "/Users/huangshifeng/Desktop/Research/AI_Clinical/痔瘡AI衛教/收案對照表_模板.xlsx"
wb.save(out_path)
print(f"Done → {out_path}")
